import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from tkinter import font as tkfont
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment


class SeatingAllocationApp:
    """Automatic Seating Allocation System — Desktop Application."""

    # ── Theme: Dark Navy / Slate ─────────────────────────────────
    THEME = {
        "BG_PRIMARY":       "#0E1419",   # Darkest navy – main background
        "BG_SECONDARY":     "#11212D",   # Dark navy – card / panel bg
        "BG_TERTIARY":      "#253745",   # Slate blue – subtle contrast
        "ACCENT_PRIMARY":   "#CCD0CF",   # Light grey – primary accent
        "ACCENT_SECONDARY": "#9BA8AB",   # Mid grey-blue – secondary accent
        "TEXT_PRIMARY":     "#CCD0CF",   # Light grey – main text
        "TEXT_SECONDARY":   "#9BA8AB",   # Mid grey-blue – hints / captions
        "TEXT_ON_ACCENT":   "#0E1419",   # Darkest navy – text on light bg
        "BORDER":           "#4A5C6A",   # Medium slate – borders
        "SUCCESS":          "#9BA8AB",   # Grey-blue – success highlight
        "ERROR":            "#C87070",   # Muted red – error highlight
        "SEAT_FILLED":      "#CCD0CF",   # Light grey – occupied seat
        "SEAT_EMPTY":       "#253745",   # Slate blue – empty seat
    }

    # ═══════════════════════════════════════════════════════════════
    #  INITIALISATION
    # ═══════════════════════════════════════════════════════════════

    def __init__(self, root):
        self.root = root
        self.root.title("Automatic Seating Allocation System")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        self.root.configure(bg=self.THEME["BG_PRIMARY"])

        # ── Data attributes (UNCHANGED) ──
        self.college_details = {}
        self.rooms = []
        self.uploaded_files = []
        self.student_data = []
        self.seating_plan = None

        # ── Internal: per-tab data display frames ──
        self._data_frames = {}

        # ── Fonts ──
        self.title_font   = tkfont.Font(family="Segoe UI", size=22, weight="bold")
        self.heading_font = tkfont.Font(family="Segoe UI", size=14, weight="bold")
        self.label_font   = tkfont.Font(family="Segoe UI", size=11)
        self.button_font  = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        self.small_font   = tkfont.Font(family="Segoe UI", size=9)
        self.data_font    = tkfont.Font(family="Segoe UI", size=10)

        # ── Header bar ──
        self._create_header()

        # ── Notebook (tab container) ──
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both", padx=20, pady=(5, 10))

        # ── Create all six tabs ──
        self.create_college_details_tab()
        self.create_room_details_tab()
        self.create_excel_upload_tab()
        self.create_generate_seating_tab()
        self.create_find_your_room_tab()
        self.create_export_tab()

        # ── Status bar ──
        self._create_status_bar()

        # ── Apply ttk theme styles ──
        self._apply_styles()

    # ═══════════════════════════════════════════════════════════════
    #  HEADER, STATUS BAR & UI HELPERS
    # ═══════════════════════════════════════════════════════════════

    def _create_header(self):
        T = self.THEME
        header = tk.Frame(self.root, bg=T["BG_SECONDARY"], height=65)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header, text="\U0001FA91  Automatic Seating Allocation System",
            font=self.title_font, bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"],
        ).pack(side="left", padx=25, pady=12)

    def _create_status_bar(self):
        T = self.THEME
        self.status_bar = tk.Frame(self.root, bg=T["BG_SECONDARY"], height=28)
        self.status_bar.pack(fill="x", side="bottom")
        self.status_bar.pack_propagate(False)
        self.status_label = tk.Label(
            self.status_bar, text="Ready", font=self.small_font,
            bg=T["BG_SECONDARY"], fg=T["TEXT_SECONDARY"], anchor="w",
        )
        self.status_label.pack(side="left", padx=15, pady=4)

    def _update_status(self):
        parts = []
        if self.rooms:
            cap = sum(r["Rows"] * r["Columns"] for r in self.rooms)
            parts.append(f"{len(self.rooms)} room(s) \u00b7 {cap} seats")
        if self.uploaded_files:
            parts.append(f"{len(self.uploaded_files)} file(s) uploaded")
        if self.student_data:
            parts.append(f"{len(self.student_data)} students loaded")
        if self.seating_plan:
            parts.append("\u2713 Seating plan generated")
        self.status_label.config(text="  \u2502  ".join(parts) if parts else "Ready")

    # ── Layout helpers ──

    def _make_tab_container(self, tab):
        T = self.THEME
        container = tk.Frame(tab, bg=T["BG_PRIMARY"])
        container.pack(expand=True, fill="both", padx=30, pady=20)
        return container

    def _make_heading(self, parent, text):
        T = self.THEME
        lbl = tk.Label(parent, text=text, font=self.heading_font,
                       bg=T["BG_PRIMARY"], fg=T["ACCENT_PRIMARY"])
        lbl.pack(anchor="w", pady=(0, 15))
        return lbl

    def _make_card(self, parent):
        T = self.THEME
        card = tk.Frame(parent, bg=T["BG_SECONDARY"], padx=30, pady=20,
                        highlightbackground=T["BORDER"], highlightthickness=1)
        card.pack(fill="x")
        return card

    def _card_label(self, card, text, row, col):
        T = self.THEME
        lbl = tk.Label(card, text=text, font=self.label_font,
                       bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"])
        lbl.grid(row=row, column=col, padx=10, pady=12, sticky="e")
        return lbl

    def _make_buttons(self, parent):
        T = self.THEME
        frame = tk.Frame(parent, bg=T["BG_PRIMARY"])
        frame.pack(pady=20)
        return frame

    def _make_data_frame(self, parent, tab):
        T = self.THEME
        frame = tk.Frame(parent, bg=T["BG_PRIMARY"])
        frame.pack(fill="x", pady=(15, 0))
        self._data_frames[tab] = frame
        return frame

    # ═══════════════════════════════════════════════════════════════
    #  TTK STYLE CONFIGURATION
    # ═══════════════════════════════════════════════════════════════

    def _apply_styles(self):
        style = ttk.Style()
        T = self.THEME

        # Notebook
        style.configure("TNotebook", background=T["BG_PRIMARY"], padding=5, borderwidth=0)
        style.configure("TNotebook.Tab",
                        font=("Segoe UI", 11, "bold"), padding=[16, 8],
                        background=T["BG_SECONDARY"], foreground=T["TEXT_SECONDARY"])
        style.map("TNotebook.Tab",
                  background=[("selected", T["BG_PRIMARY"]), ("active", T["BG_TERTIARY"])],
                  foreground=[("selected", T["ACCENT_PRIMARY"]), ("active", T["TEXT_PRIMARY"])])

        # Frames
        style.configure("TFrame", background=T["BG_PRIMARY"])

        # Labels
        style.configure("TLabel", background=T["BG_PRIMARY"],
                        font=self.label_font, foreground=T["TEXT_PRIMARY"])

        # Default button (primary look)
        style.configure("TButton", font=self.button_font, padding=[18, 10],
                        background=T["ACCENT_PRIMARY"], foreground=T["TEXT_ON_ACCENT"])
        style.map("TButton",
                  background=[("active", T["ACCENT_SECONDARY"]), ("pressed", T["ACCENT_PRIMARY"])])

        # Secondary button
        style.configure("Secondary.TButton", font=self.button_font, padding=[14, 8],
                        background=T["BG_TERTIARY"], foreground=T["TEXT_PRIMARY"])
        style.map("Secondary.TButton",
                  background=[("active", T["BORDER"]), ("pressed", T["BG_TERTIARY"])])

        # Danger button (delete)
        style.configure("Danger.TButton", font=self.button_font, padding=[14, 8],
                        background=T["ERROR"], foreground=T["TEXT_ON_ACCENT"])
        style.map("Danger.TButton",
                  background=[("active", "#A05050"), ("pressed", T["ERROR"])])

        # Entry
        style.configure("TEntry", font=self.label_font,
                        fieldbackground=T["BG_SECONDARY"], foreground=T["TEXT_PRIMARY"],
                        insertcolor=T["TEXT_PRIMARY"])

        # Combobox
        style.configure("TCombobox",
                        fieldbackground=T["BG_SECONDARY"], foreground=T["TEXT_ON_ACCENT"],
                        background=T["BG_TERTIARY"], arrowcolor=T["TEXT_PRIMARY"])
        style.map("TCombobox", fieldbackground=[("readonly", T["BG_SECONDARY"])])

        # Radiobutton
        style.configure("TRadiobutton", background=T["BG_PRIMARY"],
                        foreground=T["TEXT_PRIMARY"], font=self.label_font)
        style.map("TRadiobutton", background=[("active", T["BG_SECONDARY"])])

    # ═══════════════════════════════════════════════════════════════
    #  TAB 1 — COLLEGE DETAILS
    # ═══════════════════════════════════════════════════════════════

    def create_college_details_tab(self):
        self.college_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.college_tab, text="  College Details  ")

        container = self._make_tab_container(self.college_tab)
        self._make_heading(container, "\U0001f4cb  College & Exam Information")
        card = self._make_card(container)

        self._card_label(card, "College Name:", 0, 0)
        self.college_name_entry = ttk.Entry(card, width=40, font=self.label_font)
        self.college_name_entry.grid(row=0, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Exam Type:", 1, 0)
        self.exam_type_combo = ttk.Combobox(
            card, values=["Mid1", "Mid2", "Semester"], state="readonly", width=37)
        self.exam_type_combo.grid(row=1, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Exam Date (DD/MM/YYYY):", 2, 0)
        self.exam_date_entry = ttk.Entry(card, width=40)
        self.exam_date_entry.grid(row=2, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Exam Time (From \u2013 To):", 3, 0)
        self.exam_time_entry = ttk.Entry(card, width=40)
        self.exam_time_entry.grid(row=3, column=1, padx=10, pady=12, sticky="w")

        btns = self._make_buttons(container)
        ttk.Button(btns, text="\u2713  Submit",
                   command=self.save_college_details).pack(side="left", padx=10)
        ttk.Button(btns, text="Next  \u2192", style="Secondary.TButton",
                   command=lambda: self.notebook.select(self.room_tab)).pack(side="left", padx=10)

        self._make_data_frame(container, self.college_tab)

    # ── Logic (UNCHANGED) ─────────────────────────────────────────

    def save_college_details(self):
        college_name = self.college_name_entry.get().strip()
        exam_type = self.exam_type_combo.get()
        exam_date = self.exam_date_entry.get().strip()
        exam_time = self.exam_time_entry.get().strip()

        if not college_name or not exam_type or not exam_date or not exam_time:
            messagebox.showerror("Error", "All fields are required.")
            return

        try:
            datetime.strptime(exam_date, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use DD/MM/YYYY.")
            return

        self.college_details = {
            "College Name": college_name,
            "Exam Type": exam_type,
            "Exam Date": exam_date,
            "Exam Time": exam_time,
        }
        self.display_data_in_tab(self.college_tab, [self.college_details])
        messagebox.showinfo("Info", "College details saved successfully!")

    # ═══════════════════════════════════════════════════════════════
    #  TAB 2 — ROOM DETAILS
    # ═══════════════════════════════════════════════════════════════

    def create_room_details_tab(self):
        self.room_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.room_tab, text="  Room Details  ")

        container = self._make_tab_container(self.room_tab)
        self._make_heading(container, "\U0001f6aa  Exam Hall Configuration")
        card = self._make_card(container)

        self._card_label(card, "Room Name:", 0, 0)
        self.room_name_entry = ttk.Entry(card, width=30)
        self.room_name_entry.grid(row=0, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Rows:", 1, 0)
        self.row_entry = ttk.Entry(card, width=30)
        self.row_entry.grid(row=1, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Columns:", 2, 0)
        self.column_entry = ttk.Entry(card, width=30)
        self.column_entry.grid(row=2, column=1, padx=10, pady=12, sticky="w")

        btns = self._make_buttons(container)
        ttk.Button(btns, text="\uff0b  Add Room",
                   command=self.add_room).pack(side="left", padx=8)
        ttk.Button(btns, text="\U0001f5d1  Delete Room", style="Danger.TButton",
                   command=self.delete_room).pack(side="left", padx=8)
        ttk.Button(btns, text="Next  \u2192", style="Secondary.TButton",
                   command=lambda: self.notebook.select(self.upload_tab)).pack(side="left", padx=8)

        self._make_data_frame(container, self.room_tab)

    # ── Logic (UNCHANGED) ─────────────────────────────────────────

    def add_room(self):
        room_name = self.room_name_entry.get().strip()
        rows = self.row_entry.get().strip()
        columns = self.column_entry.get().strip()

        if not room_name or not rows or not columns:
            messagebox.showerror("Error", "All fields are required.")
            return

        try:
            rows = int(rows)
            columns = int(columns)
            if rows <= 0 or columns <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Rows and columns must be positive integers.")
            return

        room = {
            "Room Name": room_name,
            "Rows": rows,
            "Columns": columns
        }
        self.rooms.append(room)
        self.display_data_in_tab(self.room_tab, self.rooms)
        self.clear_room_entries()

    def delete_room(self):
        if self.rooms:
            self.rooms.pop()
            self.display_data_in_tab(self.room_tab, self.rooms)

    def clear_room_entries(self):
        self.room_name_entry.delete(0, tk.END)
        self.row_entry.delete(0, tk.END)
        self.column_entry.delete(0, tk.END)

    # ═══════════════════════════════════════════════════════════════
    #  TAB 3 — EXCEL UPLOAD
    # ═══════════════════════════════════════════════════════════════

    def create_excel_upload_tab(self):
        self.upload_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.upload_tab, text="  Excel Upload  ")

        container = self._make_tab_container(self.upload_tab)
        self._make_heading(container, "\U0001f4c2  Upload Student Data")
        card = self._make_card(container)

        self._card_label(card, "Branch:", 0, 0)
        self.branch_entry = ttk.Entry(card, width=30)
        self.branch_entry.grid(row=0, column=1, padx=10, pady=12, sticky="w")

        self._card_label(card, "Year:", 1, 0)
        self.year_entry = ttk.Entry(card, width=30)
        self.year_entry.grid(row=1, column=1, padx=10, pady=12, sticky="w")

        # Odd / Even semester group
        T = self.THEME
        tk.Label(card, text="Semester Group:", font=self.label_font,
                 bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"]).grid(
            row=2, column=0, padx=10, pady=12, sticky="e")

        radio_frame = tk.Frame(card, bg=T["BG_SECONDARY"])
        radio_frame.grid(row=2, column=1, padx=10, pady=12, sticky="w")

        self.branch_type = tk.StringVar(value="Odd")
        ttk.Radiobutton(radio_frame, text="  Odd  ",
                        variable=self.branch_type, value="Odd").pack(side="left", padx=(0, 20))
        ttk.Radiobutton(radio_frame, text="  Even  ",
                        variable=self.branch_type, value="Even").pack(side="left")

        btns = self._make_buttons(container)
        ttk.Button(btns, text="\U0001f4c1  Upload File",
                   command=self.upload_file).pack(side="left", padx=10)
        ttk.Button(btns, text="Next  \u2192", style="Secondary.TButton",
                   command=lambda: self.notebook.select(self.generate_tab)).pack(side="left", padx=10)

        self._make_data_frame(container, self.upload_tab)

    # ── Logic (UNCHANGED) ─────────────────────────────────────────

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                data = pd.read_excel(file_path)
                if "StudentPIN" not in data.columns:
                    messagebox.showerror("Error", "The file must contain a 'StudentPIN' column.")
                    return
                self.student_data.extend(data.to_dict(orient="records"))
                self.uploaded_files.append({
                    "Branch": self.branch_entry.get().strip(),
                    "Year": self.year_entry.get().strip(),
                    "Branch Type": self.branch_type.get(),
                    "File Path": file_path
                })
                self.display_data_in_tab(self.upload_tab, self.uploaded_files)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read the file: {str(e)}")

    # ═══════════════════════════════════════════════════════════════
    #  TAB 4 — GENERATE SEATING
    # ═══════════════════════════════════════════════════════════════

    def create_generate_seating_tab(self):
        self.generate_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.generate_tab, text="  Generate Seating  ")

        T = self.THEME
        container = tk.Frame(self.generate_tab, bg=T["BG_PRIMARY"])
        container.pack(expand=True, fill="both")

        # Centered content
        center = tk.Frame(container, bg=T["BG_PRIMARY"])
        center.place(relx=0.5, rely=0.45, anchor="center")

        tk.Label(center, text="\u26a1", font=tkfont.Font(family="Segoe UI", size=48),
                 bg=T["BG_PRIMARY"], fg=T["ACCENT_PRIMARY"]).pack(pady=(0, 10))

        tk.Label(center, text="Generate Seating Plan", font=self.heading_font,
                 bg=T["BG_PRIMARY"], fg=T["TEXT_PRIMARY"]).pack(pady=(0, 10))

        tk.Label(center,
                 text="Ensure rooms and student files are configured\nbefore generating the seating plan.",
                 font=self.small_font, bg=T["BG_PRIMARY"],
                 fg=T["TEXT_SECONDARY"], justify="center").pack(pady=(0, 25))

        ttk.Button(center, text="\U0001fa91  Generate Seating Plan",
                   command=self.generate_seating_plan).pack(pady=5)
        ttk.Button(center, text="\u2190  Back", style="Secondary.TButton",
                   command=lambda: self.notebook.select(self.upload_tab)).pack(pady=10)

    # ── Logic (UNCHANGED) ─────────────────────────────────────────

    def generate_seating_plan(self):
        if not self.rooms or not self.uploaded_files:
            messagebox.showerror("Error", "Rooms and student data are required.")
            return

        # Separate students by branch type (Odd/Even)
        odd_students, even_students = [], []
        for file_info in self.uploaded_files:
            data = pd.read_excel(file_info["File Path"])
            pins = data["StudentPIN"].tolist()
            if file_info["Branch Type"] == "Odd":
                odd_students.extend(pins)
            else:
                even_students.extend(pins)

        seating_plan = []
        odd_idx = even_idx = 0

        for room in self.rooms:
            rows, cols = room["Rows"], room["Columns"]
            seats = [["Empty" for _ in range(cols)] for _ in range(rows)]

            # Allocate odd columns to odd-branch students
            odd_cols = [col for col in range(cols) if (col + 1) % 2 != 0]
            for col in odd_cols:
                for row in range(rows):
                    if odd_idx < len(odd_students):
                        seats[row][col] = odd_students[odd_idx]
                        odd_idx += 1

            # Allocate even columns to even-branch students
            even_cols = [col for col in range(cols) if (col + 1) % 2 == 0]
            for col in even_cols:
                for row in range(rows):
                    if even_idx < len(even_students):
                        seats[row][col] = even_students[even_idx]
                        even_idx += 1

            seating_plan.append({"Room Name": room["Room Name"], "Seats": seats})

        # Dynamic adjustment for final branches to minimize empty spaces
        if odd_idx < len(odd_students) or even_idx < len(even_students):
            for room in seating_plan:
                seats = room["Seats"]
                for row in range(len(seats)):
                    for col in range(len(seats[row])):
                        if seats[row][col] == "Empty":
                            if (col + 1) % 2 != 0 and odd_idx < len(odd_students):
                                seats[row][col] = odd_students[odd_idx]
                                odd_idx += 1
                            elif (col + 1) % 2 == 0 and even_idx < len(even_students):
                                seats[row][col] = even_students[even_idx]
                                even_idx += 1

        # Handle remaining students if any
        if odd_idx < len(odd_students) or even_idx < len(even_students):
            messagebox.showwarning("Warning", "Not all students could be seated due to insufficient room capacity.")

        self.seating_plan = seating_plan
        self.show_seating_plan_popup()

    # ═══════════════════════════════════════════════════════════════
    #  TAB 5 — FIND YOUR ROOM
    # ═══════════════════════════════════════════════════════════════

    def create_find_your_room_tab(self):
        self.find_room_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.find_room_tab, text="  Find Your Room  ")

        T = self.THEME
        container = self._make_tab_container(self.find_room_tab)
        self._make_heading(container, "\U0001f50d  Student Seat Lookup")
        card = self._make_card(container)

        self._card_label(card, "Enter Student PIN:", 0, 0)
        self.student_pin_entry = ttk.Entry(card, width=30)
        self.student_pin_entry.grid(row=0, column=1, padx=10, pady=12, sticky="w")

        btns = self._make_buttons(container)
        ttk.Button(btns, text="\U0001f50d  Search",
                   command=self.find_student_room).pack(padx=10)

        # Result card
        result_card = tk.Frame(container, bg=T["BG_SECONDARY"], padx=30, pady=20,
                               highlightbackground=T["BORDER"], highlightthickness=1)
        result_card.pack(fill="x", pady=(10, 0))

        self.summary_label = tk.Label(result_card, text="", font=self.label_font,
                                      bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"])
        self.summary_label.pack(pady=5)

        self.result_label = tk.Label(
            result_card,
            text="Enter a Student PIN and click Search to find their seat.",
            font=self.label_font, bg=T["BG_SECONDARY"], fg=T["TEXT_SECONDARY"])
        self.result_label.pack(pady=5)

    # ── Logic (UNCHANGED) ─────────────────────────────────────────

    def update_branch_summary(self, event=None):
        branch = self.branch_combo.get()
        if not branch or not self.seating_plan:
            return

        branch_students = {}
        for room in self.seating_plan:
            for row in room["Seats"]:
                for pin in row:
                    if isinstance(pin, int) or (isinstance(pin, str) and pin.isdigit()):
                        if pin in self.student_data:
                            student_info = next((s for s in self.student_data if s["StudentPIN"] == pin), None)
                            if student_info and student_info.get("Branch") == branch:
                                if room["Room Name"] not in branch_students:
                                    branch_students[room["Room Name"]] = []
                                branch_students[room["Room Name"]].append(pin)

        summary_text = f"Branch: {branch}\n"
        for room, pins in branch_students.items():
            summary_text += f"{room}: {min(pins)}-{max(pins)}\n"
        self.summary_label.config(text=summary_text)

    def find_student_room(self):
        pin = self.student_pin_entry.get().strip()
        if not pin:
            messagebox.showerror("Error", "Please enter a Student PIN.")
            return

        for room in self.seating_plan:
            for row_idx, row in enumerate(room["Seats"]):
                for col_idx, seat in enumerate(row):
                    if str(seat) == pin:
                        self.result_label.config(text=f"Room: {room['Room Name']}, Row: {row_idx + 1}, Column: {col_idx + 1}")
                        return

        self.result_label.config(text="Student PIN not found.")

    # ═══════════════════════════════════════════════════════════════
    #  TAB 6 — EXPORT
    # ═══════════════════════════════════════════════════════════════

    def create_export_tab(self):
        self.export_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.export_tab, text="  Export  ")

        T = self.THEME
        container = tk.Frame(self.export_tab, bg=T["BG_PRIMARY"])
        container.pack(expand=True, fill="both")

        center = tk.Frame(container, bg=T["BG_PRIMARY"])
        center.place(relx=0.5, rely=0.45, anchor="center")

        tk.Label(center, text="\U0001f4e4", font=tkfont.Font(family="Segoe UI", size=48),
                 bg=T["BG_PRIMARY"], fg=T["ACCENT_PRIMARY"]).pack(pady=(0, 10))

        tk.Label(center, text="Export Seating Plan", font=self.heading_font,
                 bg=T["BG_PRIMARY"], fg=T["TEXT_PRIMARY"]).pack(pady=(0, 10))

        tk.Label(center,
                 text="Export a formatted Excel workbook with one sheet per room,\ncollege info, seating grid, and attendance summary.",
                 font=self.small_font, bg=T["BG_PRIMARY"],
                 fg=T["TEXT_SECONDARY"], justify="center").pack(pady=(0, 25))

        ttk.Button(center, text="\U0001f4ca  Export to Excel",
                   command=self.export_to_excel).pack(pady=5)
        ttk.Button(center, text="\u2190  Back", style="Secondary.TButton",
                   command=lambda: self.notebook.select(self.generate_tab)).pack(pady=10)

    # ═══════════════════════════════════════════════════════════════
    #  DATA TABLE DISPLAY (redesigned styling, same data flow)
    # ═══════════════════════════════════════════════════════════════

    def display_data_in_tab(self, tab, data):
        T = self.THEME

        # Use dedicated data frame if available (new card-based layout)
        data_frame = self._data_frames.get(tab)
        if data_frame:
            for widget in list(data_frame.winfo_children()):
                widget.destroy()
        else:
            # Fallback: original row-based cleanup for compatibility
            for widget in list(tab.winfo_children()):
                try:
                    row_info = widget.grid_info().get("row", 0)
                except Exception:
                    row_info = 0
                if isinstance(widget, ttk.Label) and widget["text"] != "" and row_info >= 6:
                    widget.destroy()
            data_frame = tab

        if not data:
            self._update_status()
            return

        # Determine headers
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            headers = ["Value"]

        # Header row — accent background
        for col, header in enumerate(headers):
            lbl = tk.Label(data_frame, text=header, font=self.label_font,
                           bg=T["ACCENT_PRIMARY"], fg=T["TEXT_ON_ACCENT"],
                           padx=12, pady=8, relief="flat")
            lbl.grid(row=0, column=col, padx=1, pady=(0, 1), sticky="ew")

        # Data rows — zebra striping
        for row_idx, item in enumerate(data):
            bg = T["BG_SECONDARY"] if row_idx % 2 == 0 else T["BG_TERTIARY"]
            for col_idx, key in enumerate(headers):
                value = item[key] if isinstance(item, dict) else item
                lbl = tk.Label(data_frame, text=str(value), font=self.data_font,
                               bg=bg, fg=T["TEXT_PRIMARY"], padx=12, pady=6, relief="flat")
                lbl.grid(row=1 + row_idx, column=col_idx, padx=1, pady=0, sticky="ew")

        self._update_status()

    # ═══════════════════════════════════════════════════════════════
    #  SEATING PLAN POPUP (redesigned visuals)
    # ═══════════════════════════════════════════════════════════════

    def show_seating_plan_popup(self):
        T = self.THEME

        seating_popup = tk.Toplevel(self.root)
        seating_popup.title("Seating Plan")
        seating_popup.geometry("900x650")
        seating_popup.configure(bg=T["BG_PRIMARY"])

        # Canvas with both scrollbars
        outer = tk.Frame(seating_popup, bg=T["BG_PRIMARY"])
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, bg=T["BG_PRIMARY"], highlightthickness=0)
        v_scroll = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        h_scroll = tk.Scrollbar(outer, orient="horizontal", command=canvas.xview)
        canvas.config(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)

        frame = tk.Frame(canvas, bg=T["BG_PRIMARY"])
        canvas.create_window((0, 0), window=frame, anchor="nw")

        # College info card
        info_card = tk.Frame(frame, bg=T["BG_SECONDARY"], padx=25, pady=15,
                             highlightbackground=T["BORDER"], highlightthickness=1)
        info_card.pack(fill="x", padx=20, pady=(15, 10))

        college_info = (
            f"College Name: {self.college_details['College Name']}    \u2502    "
            f"Exam Type: {self.college_details['Exam Type']}    \u2502    "
            f"Exam Date: {self.college_details['Exam Date']}    \u2502    "
            f"Exam Time: {self.college_details['Exam Time']}"
        )
        tk.Label(info_card, text=college_info, font=self.label_font,
                 bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"],
                 wraplength=820, justify="left").pack(anchor="w")

        total_seats = 0
        for room in self.seating_plan:
            # Divider line
            tk.Frame(frame, bg=T["BORDER"], height=1).pack(fill="x", padx=20, pady=15)

            # Room heading
            tk.Label(frame, text=f"\U0001f6aa  {room['Room Name']}", font=self.heading_font,
                     bg=T["BG_PRIMARY"], fg=T["ACCENT_PRIMARY"]).pack(anchor="w", padx=25, pady=(0, 10))

            seats = room["Seats"]
            total_seats += len(seats) * len(seats[0])

            # Seat grid
            table_frame = tk.Frame(frame, bg=T["BG_PRIMARY"])
            table_frame.pack(pady=5, padx=25)

            for row_idx, row in enumerate(seats):
                for col_idx, seat in enumerate(row):
                    is_empty = (seat == "Empty")
                    bg = T["SEAT_EMPTY"] if is_empty else T["SEAT_FILLED"]
                    fg = T["TEXT_SECONDARY"] if is_empty else T["TEXT_ON_ACCENT"]
                    text = "\u2014" if is_empty else str(seat)
                    cell_font = self.small_font if is_empty else self.data_font

                    lbl = tk.Label(table_frame, text=text, width=12, relief="flat",
                                   padx=5, pady=4, bg=bg, fg=fg, font=cell_font)
                    lbl.grid(row=row_idx, column=col_idx, padx=2, pady=2)

            # Stats row
            total_allotted = sum(1 for row in seats for seat in row if seat != "Empty")
            total_present = 0  # Placeholder for actual attendance tracking
            total_absent = 0

            stats_frame = tk.Frame(frame, bg=T["BG_SECONDARY"], padx=20, pady=10,
                                   highlightbackground=T["BORDER"], highlightthickness=1)
            stats_frame.pack(fill="x", padx=25, pady=(10, 5))

            for i, (stat_label, stat_val) in enumerate([
                (f"Allotted in {room['Room Name']}", total_allotted),
                ("Present", total_present),
                ("Absent", total_absent),
            ]):
                tk.Label(stats_frame, text=f"{stat_label}: {stat_val}", font=self.data_font,
                         bg=T["BG_SECONDARY"], fg=T["TEXT_PRIMARY"]).grid(
                    row=0, column=i, padx=15)

            tk.Label(frame,
                     text="Signature of the Invigilator:  ____________________________",
                     font=self.label_font, bg=T["BG_PRIMARY"],
                     fg=T["TEXT_SECONDARY"]).pack(anchor="w", padx=25, pady=(5, 0))

        # Total footer
        tk.Frame(frame, bg=T["ACCENT_PRIMARY"], height=2).pack(fill="x", padx=20, pady=15)
        tk.Label(frame, text=f"Total Seats: {total_seats}", font=self.heading_font,
                 bg=T["BG_PRIMARY"], fg=T["ACCENT_PRIMARY"]).pack(pady=(0, 20))

        # Update scroll region
        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        self._update_status()

    # ═══════════════════════════════════════════════════════════════
    #  EXPORT TO EXCEL (Logic UNCHANGED — indentation corrected)
    # ═══════════════════════════════════════════════════════════════

    def export_to_excel(self):
        if not self.seating_plan:
            messagebox.showerror("Error", "No seating plan to export.")
            return

        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                wb = Workbook()
                for room in self.seating_plan:
                    ws = wb.create_sheet(title=room["Room Name"])
                    ws.append(["College Name", self.college_details["College Name"]])
                    ws.append(["Exam Type", self.college_details["Exam Type"]])
                    ws.append(["Exam Date", self.college_details["Exam Date"]])
                    ws.append(["Exam Time", self.college_details["Exam Time"]])
                    ws.append([])
                    ws.append(["Room Name", room["Room Name"]])
                    ws.append([])

                    # Write seating plan
                    for row_idx, row in enumerate(room["Seats"]):
                        ws.append([f"Row {row_idx + 1}"] + row)

                    # Add labels
                    total_allotted = sum(1 for row in room["Seats"] for seat in row if seat != "Empty")
                    total_present = 0  # Placeholder for actual attendance tracking
                    total_absent = 0

                    ws.append([])
                    ws.append(["Total Allotted", total_allotted])
                    ws.append(["Total Present", total_present])
                    ws.append(["Total Absent", total_absent])
                    ws.append(["Signature of the Invigilator", "___________________"])

                    # Adjust column widths dynamically
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
                        for cell in col:
                            try:
                                # Calculate the length of the cell content
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                        # Set column width to fit the longest content
                        adjusted_width = (max_length + 2) * 1.2  # Add padding and scaling factor
                        ws.column_dimensions[column].width = adjusted_width

                # Remove the default sheet created by openpyxl
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

                # Save the workbook
                wb.save(file_path)
                messagebox.showinfo("Success", "Seating plan exported successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = SeatingAllocationApp(root)
    root.mainloop()